import openpyxl
from teleport.teleport_sim import TeleportSimulator
from ai.engine import SERAIEngine
from datetime import datetime

# ----------------------------
# STEP 1: Load Aura.xlsx
# ----------------------------
WORKBOOK_PATH = "extensions/Aura.xlsx"
engine = SERAIEngine(WORKBOOK_PATH)
print("Sheets in Aura.xlsx:", engine.list_sheets())

# ----------------------------
# STEP 2: Add a new STEM module
# ----------------------------
wb = openpyxl.load_workbook(WORKBOOK_PATH)
if "Quantum_Math" not in wb.sheetnames:
    wb.create_sheet("Quantum_Math")
sheet = wb["Quantum_Math"]
sheet.append(["Topic", "Formula", "Notes"])
sheet.append(["Quantum Entanglement", "ψ = α|0> + β|1>", "Sample entry"])
wb.save(WORKBOOK_PATH)
print("Added 'Quantum_Math' sheet with sample data.")

# ----------------------------
# STEP 3: Add teleportation log sheet (extension)
# ----------------------------
if "Teleport_Log" not in wb.sheetnames:
    wb.create_sheet("Teleport_Log")
sheet = wb["Teleport_Log"]
sheet.append(["Time", "TP Stage", "Outcome", "Notes"])
wb.save(WORKBOOK_PATH)
print("Added 'Teleport_Log' sheet.")

# ----------------------------
# STEP 4: Run basic teleport simulation
# ----------------------------
simulator = TeleportSimulator(engine)
sim_results = simulator.run_basic_simulation()
print("Simulation Results:")
for stage, outcome in sim_results.items():
    print(f"{stage}: {outcome}")

# ----------------------------
# STEP 5: Log results automatically
# ----------------------------
sheet = wb["Teleport_Log"]
for stage, outcome in sim_results.items():
    sheet.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), stage, outcome, "Auto-simulated"])
wb.save(WORKBOOK_PATH)
print("Logged simulation results to 'Teleport_Log'.")

# ----------------------------
# STEP 6: Verify logs
# ----------------------------
for row in wb["Teleport_Log"].iter_rows(values_only=True):
    print(row)
